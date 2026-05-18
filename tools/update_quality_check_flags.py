#!/usr/bin/env python3
"""Refresh reproducible `quality_check` flags for FOMC press-conference metadata.

Method:
- fixed random seed = 42
- 5% sample of the press-conference corpus
- sample size rounded up with `ceil`

This keeps the manual quality-control sample reproducible and visible directly
in the press-conference metadata files only.
"""

from __future__ import annotations

import math
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
METADATA_DIR = ROOT / "data" / "metadata"
SEED = 42
SAMPLE_SHARE = 0.05

CONFIG = [
    {
        "csv": "fomc_press_conferences.csv",
        "group_col": "document_type",
        "label": "FOMC press conferences",
    },
]

LINK_FONT = Font(color="0563C1", underline="single")
URL_COLUMNS = ("speech_url", "source_url")
PATH_COLUMNS = ("text_path",)


def _sorted_with_original_index(df: pd.DataFrame) -> pd.DataFrame:
    sort_cols = [
        col
        for col in ["meeting_date", "document_date", "date", "text_path", "title"]
        if col in df.columns
    ]
    if not sort_cols:
        return df.reset_index(names="_original_index")
    return df.sort_values(sort_cols, kind="mergesort").reset_index(names="_original_index")


def _sample_size(n_rows: int) -> int:
    if n_rows <= 0:
        return 0
    return min(n_rows, max(1, math.ceil(n_rows * SAMPLE_SHARE)))


def build_quality_flags(df: pd.DataFrame, group_col: str | None) -> pd.Series:
    flags = pd.Series(False, index=df.index, dtype=bool)
    if df.empty:
        return flags

    ordered = _sorted_with_original_index(df)

    if group_col and group_col in ordered.columns:
        for _, group in ordered.groupby(group_col, dropna=False, sort=True):
            n_sample = _sample_size(len(group))
            chosen = group.sample(n=n_sample, random_state=SEED)["_original_index"]
            flags.loc[chosen.to_list()] = True
    else:
        n_sample = _sample_size(len(ordered))
        chosen = ordered.sample(n=n_sample, random_state=SEED)["_original_index"]
        flags.loc[chosen.to_list()] = True

    return flags


def _to_file_uri(rel_path: str) -> str | None:
    if not rel_path or pd.isna(rel_path):
        return None
    normalized = str(rel_path).replace("\\", "/")
    try:
        return (ROOT / normalized).resolve().as_uri()
    except Exception:
        return None


def add_hyperlinks(xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    url_col_indices = [headers.index(col) + 1 for col in URL_COLUMNS if col in headers]
    path_col_indices = [headers.index(col) + 1 for col in PATH_COLUMNS if col in headers]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in url_col_indices:
            cell = row[col_idx - 1]
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = LINK_FONT

        for col_idx in path_col_indices:
            cell = row[col_idx - 1]
            if cell.value:
                uri = _to_file_uri(str(cell.value))
                if uri:
                    cell.hyperlink = uri
                    cell.font = LINK_FONT

    wb.save(xlsx_path)


def write_excel(df: pd.DataFrame, xlsx_path: Path) -> Path:
    try:
        df.to_excel(xlsx_path, index=False)
        add_hyperlinks(xlsx_path)
        return xlsx_path
    except PermissionError:
        fallback = xlsx_path.with_name(f"{xlsx_path.stem}_updated{xlsx_path.suffix}")
        df.to_excel(fallback, index=False)
        add_hyperlinks(fallback)
        print(f"WARNING: {xlsx_path.name} is locked; wrote updated copy to {fallback.name}")
        return fallback


def print_summary(df: pd.DataFrame, label: str, group_col: str | None) -> None:
    total_flagged = int(df["quality_check"].sum())
    print(f"\n{label}: {total_flagged} / {len(df)} flagged for review")

    if group_col and group_col in df.columns:
        grouped = (
            df.groupby(group_col, dropna=False)["quality_check"]
            .agg(total="size", flagged="sum")
            .reset_index()
        )
        for _, row in grouped.iterrows():
            pct = (float(row["flagged"]) / float(row["total"]) * 100.0) if row["total"] else 0.0
            print(f"  - {row[group_col]}: {int(row['flagged'])} / {int(row['total'])} ({pct:.2f}%)")


def main() -> None:
    print(f"Refreshing quality-check flags with seed={SEED} and sample_share={SAMPLE_SHARE:.0%}")

    for cfg in CONFIG:
        csv_path = METADATA_DIR / cfg["csv"]
        xlsx_path = csv_path.with_suffix(".xlsx")

        df = pd.read_csv(csv_path)
        df["quality_check"] = build_quality_flags(df, cfg["group_col"])

        df.to_csv(csv_path, index=False)
        written_xlsx = write_excel(df, xlsx_path)

        print_summary(df, cfg["label"], cfg["group_col"])
        print(f"  CSV : {csv_path.name}")
        print(f"  XLSX: {written_xlsx.name}")


if __name__ == "__main__":
    main()
