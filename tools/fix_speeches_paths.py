"""Fix text_path in fed_speeches.csv and generate Excel with hyperlinks."""
import pandas as pd
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data/metadata/fed_speeches.csv"
XLSX_PATH = ROOT / "data/metadata/fed_speeches.xlsx"

speeches = pd.read_csv(CSV_PATH)

def fix_path(p: str) -> str:
    p = p.replace("\\", "/")
    # If path is just "fed_speeches/filename.txt", expand to full relative path
    if p.startswith("fed_speeches/"):
        p = "data/processed/fed_texts/fed_speeches/" + p.split("/")[-1]
    return p

speeches["text_path"] = speeches["text_path"].apply(fix_path)
speeches.to_csv(CSV_PATH, index=False)

print("Fixed text_path samples:")
for p in speeches["text_path"].head(5):
    print(" ", p)

# --- Excel with clickable hyperlinks ---
try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "fed_speeches"

    ws.append(speeches.columns.tolist())
    for cell in ws[1]:
        cell.font = Font(bold=True)

    col_indices = {col: i + 1 for i, col in enumerate(speeches.columns)}

    for _, row in speeches.iterrows():
        ws.append(row.tolist())
        excel_row = ws.max_row

        # speech_url → clickable hyperlink
        url_col = col_indices["speech_url"]
        url_val = str(row["speech_url"])
        if url_val.startswith("http"):
            cell = ws.cell(row=excel_row, column=url_col)
            cell.hyperlink = url_val
            cell.value = url_val
            cell.font = Font(color="0563C1", underline="single")

        # text_path → clickable file:// hyperlink
        tp_col = col_indices["text_path"]
        rel = str(row["text_path"])
        abs_path = (ROOT / rel).resolve()
        file_uri = abs_path.as_uri()
        cell = ws.cell(row=excel_row, column=tp_col)
        cell.hyperlink = file_uri
        cell.value = rel
        cell.font = Font(color="0563C1", underline="single")

    for col_idx, col_name in enumerate(speeches.columns, start=1):
        max_len = max(len(str(col_name)), speeches[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(XLSX_PATH)
    print(f"\nExcel file with hyperlinks saved to: {XLSX_PATH}")

except ImportError:
    print("\nopenpyxl not installed — skipping Excel export.")
