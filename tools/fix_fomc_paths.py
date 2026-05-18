"""Fix text_path in fomc_minutes_statements.csv and generate Excel with hyperlinks."""
import pandas as pd
from pathlib import Path
import re

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data/metadata/fomc_minutes_statements.csv"
XLSX_PATH = ROOT / "data/metadata/fomc_minutes_statements.xlsx"

fomc = pd.read_csv(CSV_PATH)

def fix_path(p: str) -> str:
    p = p.replace("\\", "/")
    for old in [
        "data/raw/fed_texts/fomc_materials_structured/",
        "data/raw/fed_texts/fomc_minutes_statements/",
    ]:
        if old in p:
            p = p.replace(old, "data/processed/fed_texts/fomc_minutes_statements/")
    return p

fomc["text_path"] = fomc["text_path"].apply(fix_path)
fomc.to_csv(CSV_PATH, index=False)

print("Fixed text_path samples:")
for p in fomc["text_path"].head(5):
    print(" ", p)

# --- Excel with clickable hyperlinks ---
try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "fomc_minutes_statements"

    # Header row
    ws.append(fomc.columns.tolist())
    for cell in ws[1]:
        cell.font = Font(bold=True)

    url_cols = {"source_url", "text_path"}
    col_indices = {col: i + 1 for i, col in enumerate(fomc.columns)}

    for _, row in fomc.iterrows():
        ws.append(row.tolist())
        excel_row = ws.max_row

        # Make source_url a clickable hyperlink
        src_col = col_indices["source_url"]
        src_val = str(row["source_url"])
        if src_val.startswith("http"):
            cell = ws.cell(row=excel_row, column=src_col)
            cell.hyperlink = src_val
            cell.value = src_val
            cell.font = Font(color="0563C1", underline="single")

        # Make text_path a clickable file:// hyperlink (absolute)
        tp_col = col_indices["text_path"]
        rel = str(row["text_path"])
        abs_path = (ROOT / rel).resolve()
        file_uri = abs_path.as_uri()
        cell = ws.cell(row=excel_row, column=tp_col)
        cell.hyperlink = file_uri
        cell.value = rel          # display the short relative path
        cell.font = Font(color="0563C1", underline="single")

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(fomc.columns, start=1):
        max_len = max(len(str(col_name)), fomc[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(XLSX_PATH)
    print(f"\nExcel file with hyperlinks saved to: {XLSX_PATH}")

except ImportError:
    print("\nopenpyxl not installed — skipping Excel export. Run: pip install openpyxl")
