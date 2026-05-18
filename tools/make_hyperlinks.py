import pandas as pd
import openpyxl
from openpyxl.styles import Font

BASE = "c:/Users/lione/OneDrive/HSG_Unterlagen/6. Semester FGV/Bachelorarbeit/FED Data/Code"
xlsx_path = BASE + "/data/metadata/fed_speeches.xlsx"
csv_path  = BASE + "/data/metadata/fed_speeches.csv"

# Rebuild xlsx from CSV (source of truth)
df = pd.read_csv(csv_path)
df.to_excel(xlsx_path, index=False)

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

headers = [cell.value for cell in ws[1]]
url_col  = headers.index("speech_url") + 1
path_col = headers.index("text_path") + 1

link_font = Font(color="0563C1", underline="single")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # speech_url -> clickable http link
    url_cell = row[url_col - 1]
    if url_cell.value:
        url_cell.hyperlink = url_cell.value
        url_cell.font = link_font

    # text_path -> absolute file:// URI
    path_cell = row[path_col - 1]
    if path_cell.value:
        rel = path_cell.value.replace("\\", "/")
        abs_uri = "file:///" + BASE.lstrip("/") + "/" + rel.lstrip("/")
        path_cell.hyperlink = abs_uri
        path_cell.font = link_font

wb.save(xlsx_path)
print(f"Done. Hyperlinks written for {ws.max_row - 1} rows.")
